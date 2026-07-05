"""
Robô — baixa o relatório "dinheiro em conta" (SETTLEMENT) do Mercado Pago e
guarda em 'settlement_mp'. É o relatório que TEM o EXTERNAL_REFERENCE (order_id
+ shipping_id), o elo que faltava pra casar o recebido com as vendas.

Feito no MESMO molde do sync_repasses_mp.py (que já funciona):
  1) renova o token do ML (que também acessa o Mercado Pago)
  2) padroniza as colunas da conta (config) — inclui EXTERNAL_REFERENCE e o cupom
  3) manda GERAR o relatório do período (POST) — assíncrono, entra "DELAYED"
  4) espera ficar pronto (o relatório fica pronto quando o file_name aparece)
  5) baixa o CSV e lê pelo CABEÇALHO (colunas por nome)
  6) grava em 'settlement_mp' sem duplicar (hash + upsert)

Família da API: /v1/account/settlement_report (diferente do release_report).
Lista pelo /search (retorna {"results":[...]}).

MODOS (env MODO):
  sonda  -> só diagnóstico (mostra config e /search). Não gera, não grava.
  gerar  -> só dispara o pedido do relatório e sai (não espera).
  baixar -> recolhe o relatório pronto mais recente, mostra amostra e GRAVA.
  cheio  -> dispara todas as contas, espera até ~30 min, baixa e GRAVA. (produção/backfill)

Env: ML_CLIENT_ID, ML_CLIENT_SECRET, SUPABASE_URL, SUPABASE_KEY
     MODO (padrão sonda) | DIAS (padrão 45) | BEGIN/END (ISO fixo, p/ backfill)
     ONLY_SELLER (roda só 1 conta)
"""

import os
import io
import csv
import json
import time
import hashlib
import requests
from datetime import datetime, timedelta, timezone
from supabase import create_client

CLIENT_ID     = os.environ["ML_CLIENT_ID"]
CLIENT_SECRET = os.environ["ML_CLIENT_SECRET"]
SEED_REFRESH  = os.environ.get("ML_REFRESH_TOKEN", "")
SUPABASE_URL  = os.environ["SUPABASE_URL"]
SUPABASE_KEY  = os.environ["SUPABASE_KEY"]

MODO        = os.environ.get("MODO", "sonda").strip().lower()
DIAS        = int(os.environ.get("DIAS", "45"))
BEGIN_FIX   = os.environ.get("BEGIN", "").strip()
END_FIX     = os.environ.get("END", "").strip()
ONLY_SELLER = os.environ.get("ONLY_SELLER", "").strip()

ML     = "https://api.mercadolibre.com"
MP     = "https://api.mercadopago.com"
REPORT = "/v1/account/settlement_report"

# colunas EXATAMENTE como no arquivo da CF que validamos (cupom vem em OPERATION_TAGS).
COLUNAS = ["EXTERNAL_REFERENCE", "SOURCE_ID", "TRANSACTION_AMOUNT", "TRANSACTION_DATE",
           "FEE_AMOUNT", "SETTLEMENT_NET_AMOUNT", "SETTLEMENT_DATE",
           "TAXES_AMOUNT", "TAX_DETAIL", "TAXES_DISAGGREGATED", "DESCRIPTION",
           "OPERATION_TAGS", "SUB_UNIT", "PRODUCT_SKU", "SALE_DETAIL"]

# cabeçalho do CSV -> coluna da tabela
MAPA = {
    "EXTERNAL_REFERENCE":    "external_reference",
    "SOURCE_ID":             "source_id",
    "TRANSACTION_AMOUNT":    "transaction_amount",
    "TRANSACTION_DATE":      "transaction_date",
    "FEE_AMOUNT":            "fee_amount",
    "SETTLEMENT_NET_AMOUNT": "net_amount",
    "SETTLEMENT_DATE":       "settlement_date",
    "PRODUCT_SKU":           "sku",
    "SALE_DETAIL":           "sale_detail",
    "OPERATION_TAGS":        "operation_tags",
}
NUMERICAS = {"transaction_amount", "fee_amount", "net_amount"}
DATAS     = {"transaction_date", "settlement_date"}

sb = create_client(SUPABASE_URL, SUPABASE_KEY)


def renovar_token(refresh):
    r = requests.post(ML + "/oauth/token", data={
        "grant_type": "refresh_token", "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET, "refresh_token": refresh}, timeout=30)
    d = r.json()
    if "access_token" not in d:
        raise RuntimeError("Falha ao renovar token: " + str(d)[:200])
    return d


def lista_contas():
    res = sb.table("contas").select("seller_id, apelido, refresh_token").execute()
    tk = [(str(c["seller_id"]), c.get("apelido"), c["refresh_token"])
          for c in (res.data or []) if c.get("refresh_token")]
    if not tk and SEED_REFRESH:
        tk = [(None, "(seed)", SEED_REFRESH)]
    if ONLY_SELLER:
        tk = [t for t in tk if t[0] == ONLY_SELLER]
    return tk


def H(access):
    return {"Authorization": "Bearer " + access}


def Hj(access):
    return {"Authorization": "Bearer " + access, "Content-Type": "application/json"}


# ---------- config das colunas ----------
def garantir_config(access, seller_id):
    body = {
        "file_name_prefix": f"settlement-report-{seller_id}",
        "include_withdraw": False,
        "show_chargeback_cancel": True,
        "scheduled": False,
        "frequency": {"hour": 0, "type": "daily", "value": 1},
        "separator": ",",
        "columns": [{"key": k} for k in COLUNAS],
    }
    r = requests.post(MP + REPORT + "/config", headers=Hj(access),
                      data=json.dumps(body), timeout=30)
    if r.status_code not in (200, 201):
        r = requests.put(MP + REPORT + "/config", headers=Hj(access),
                         data=json.dumps(body), timeout=30)
    print(f"  config -> {r.status_code}")
    return r.status_code in (200, 201)


# ---------- gerar / listar / achar ----------
def gerar(access, ini_s, fim_s):
    corpo = json.dumps({"begin_date": ini_s, "end_date": fim_s})
    g = requests.post(MP + REPORT, headers=Hj(access), data=corpo, timeout=30)
    ok = g.status_code in (200, 202)
    print(f"  gerar {ini_s[:10]}..{fim_s[:10]} -> {g.status_code} {g.text[:120]}")
    return ok


def lista_relatorios(access):
    r = requests.get(MP + REPORT + "/search?limit=100", headers=H(access), timeout=30)
    if r.status_code != 200:
        return []
    try:
        return (r.json() or {}).get("results") or []
    except Exception:
        return []


def achar_pronto(access, corte, alvo_ini):
    """Pronto = file_name preenchido. Escolhe o que COBRE a faixa pedida:
       termina recente (end_date >= corte) e começa lá atrás (begin_date <= alvo_ini)."""
    cand = [x for x in lista_relatorios(access)
            if x.get("file_name")
            and (str(x.get("end_date", ""))[:10]  >= corte)
            and (str(x.get("begin_date", ""))[:10] <= alvo_ini)]
    if not cand:
        return None
    cand.sort(key=lambda x: x.get("date_created", ""), reverse=True)
    cand.sort(key=lambda x: x.get("begin_date", ""))
    return cand[0]


def achar_mais_recente(access):
    """Pra modo 'baixar': o relatório pronto mais novo, sem exigir faixa."""
    cand = [x for x in lista_relatorios(access) if x.get("file_name")]
    if not cand:
        return None
    cand.sort(key=lambda x: x.get("date_created", ""), reverse=True)
    return cand[0]


# ---------- baixar / parsear / gravar ----------
def baixar(access, file_name):
    r = requests.get(MP + REPORT + "/" + file_name, headers=H(access), timeout=120)
    if r.status_code != 200:
        print(f"  download {file_name} -> {r.status_code} {r.text[:150]}")
        return None
    return r.content                                   # bytes (pode ser CSV ou XLSX)


def num(v):
    if v is None:
        return None
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if s == "":
        return None
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except Exception:
        return None


def data(v):
    if not v:
        return None
    s = str(v).strip()
    return s[:10] if len(s) >= 10 else None


def ler_registros(file_name, conteudo):
    """Lê o arquivo (CSV OU XLSX) e devolve dicts com cabeçalho em MAIÚSCULO."""
    if (file_name or "").lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h or "").strip().upper() for h in rows[0]]
        regs = []
        for r in rows[1:]:
            regs.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
        return regs
    # CSV
    texto = conteudo.decode("utf-8", errors="replace")
    amostra = texto[:2000]
    try:
        sep = csv.Sniffer().sniff(amostra, delimiters=";,").delimiter
    except Exception:
        sep = ";" if amostra.count(";") >= amostra.count(",") else ","
    rd = csv.DictReader(io.StringIO(texto), delimiter=sep)
    return [{(k or "").strip().upper(): v for k, v in row.items()} for row in rd]


def montar_linhas(regs, seller_id):
    linhas = []
    for rr in regs:
        out = {"seller_id": seller_id}
        for col_csv, col_tab in MAPA.items():
            val = rr.get(col_csv)
            if col_tab in NUMERICAS:
                out[col_tab] = num(val)
            elif col_tab in DATAS:
                out[col_tab] = data(val)
            else:
                out[col_tab] = (str(val).strip() if val not in (None, "") else None)
        # só o que assentou de fato (tem data de liberação e valor líquido)
        if not (out.get("settlement_date") and out.get("net_amount") is not None):
            continue
        base = "|".join(str(out.get(c) or "") for c in
                        ("seller_id", "source_id", "external_reference",
                         "settlement_date", "net_amount", "transaction_amount", "fee_amount"))
        out["hash"] = hashlib.md5(base.encode()).hexdigest()
        linhas.append(out)
    return linhas


def gravar(linhas):
    n = 0
    for i in range(0, len(linhas), 200):
        lote = linhas[i:i + 200]
        try:
            sb.table("settlement_mp").upsert(lote, on_conflict="hash").execute()
            n += len(lote)
        except Exception as e:
            print("  erro ao gravar lote:", str(e)[:120])
    return n


def mostrar(regs, linhas):
    print(f"  registros no arquivo: {len(regs)}")
    if regs:
        print("  COLUNAS:", list(regs[0].keys()))
        cab = list(regs[0].keys())[:12]
        print("  1ª linha crua:", {k: regs[0].get(k) for k in cab})
    print(f"  PARSEADAS (settled): {len(linhas)}")
    for r in linhas[:3]:
        print("    ->", {k: r.get(k) for k in
              ("external_reference", "source_id", "transaction_amount",
               "fee_amount", "net_amount", "settlement_date", "operation_tags")})


# ---------- janela ----------
def janela():
    if BEGIN_FIX and END_FIX:
        return BEGIN_FIX + "T00:00:00Z", END_FIX + "T23:59:59Z"
    fim = datetime.now(timezone.utc)
    ini = fim - timedelta(days=DIAS)
    return ini.strftime("%Y-%m-%dT%H:%M:%SZ"), fim.strftime("%Y-%m-%dT%H:%M:%SZ")


# ---------- modos ----------
def sonda(access, seller_id):
    print(f"\n=== [{seller_id}] SONDA ===")
    r = requests.get(MP + "/users/me", headers=H(access), timeout=20)
    print(f"  /users/me -> {r.status_code} {r.json().get('nickname') if r.status_code==200 else r.text[:120]}")
    rels = lista_relatorios(access)
    print(f"  /search -> {len(rels)} relatório(s). Últimos:")
    for x in rels[:5]:
        print(f"    id={x.get('id')} status={x.get('status')} file={x.get('file_name') or '-'} "
              f"faixa={str(x.get('begin_date',''))[:10]}..{str(x.get('end_date',''))[:10]}")


def modo_gerar(access, seller_id):
    print(f"\n=== [{seller_id}] GERAR ===")
    garantir_config(access, seller_id)
    ini_s, fim_s = janela()
    gerar(access, ini_s, fim_s)
    print("  pedido enviado. Recolha depois com MODO=baixar (ou cheio).")


def modo_baixar(access, seller_id, gravar_db=True):
    print(f"\n=== [{seller_id}] BAIXAR ===")
    rel = achar_mais_recente(access)
    if not rel:
        print("  nenhum relatório pronto ainda (file_name vazio). Rode MODO=gerar e espere uns minutos.")
        return
    print(f"  pronto: id={rel.get('id')} file={rel.get('file_name')} "
          f"faixa={str(rel.get('begin_date',''))[:10]}..{str(rel.get('end_date',''))[:10]}")
    conteudo = baixar(access, rel["file_name"])
    if not conteudo:
        return
    regs = ler_registros(rel["file_name"], conteudo)
    linhas = montar_linhas(regs, seller_id)
    mostrar(regs, linhas)
    if gravar_db:
        n = gravar(linhas)
        print(f"  gravadas/atualizadas: {n} linhas em settlement_mp.")


def main():
    contas = lista_contas()
    if not contas:
        print("Nenhuma conta com refresh_token em `contas`.")
        return
    print(f"MODO={MODO} | contas={[c[0] for c in contas]} | DIAS={DIAS}")

    # autentica todas as contas
    sess = []
    for sid, apelido, refresh in contas:
        try:
            d = renovar_token(refresh)
        except Exception as e:
            print(f"[{sid}] falha ao renovar token: {e}")
            continue
        access = d["access_token"]
        sid_real = str(d.get("user_id") or sid or "")
        try:
            sb.table("contas").upsert(
                {"seller_id": sid_real, "refresh_token": d.get("refresh_token", refresh)},
                on_conflict="seller_id").execute()
        except Exception:
            pass
        sess.append((sid_real, apelido, access))

    # MODO CHEIO = igual ao robô de repasses: gera TODAS, espera ~30 min, baixa e grava
    if MODO == "cheio":
        ini_s, fim_s = janela()
        for sid, apelido, access in sess:
            print(f"\n[gerar] {apelido or sid}")
            garantir_config(access, sid)
            gerar(access, ini_s, fim_s)
        # faixa alvo pra reconhecer o relatório certo
        ini_d = ini_s[:10]
        fim_d = fim_s[:10]
        corte    = (datetime.strptime(fim_d, "%Y-%m-%d") - timedelta(days=2)).strftime("%Y-%m-%d")
        alvo_ini = (datetime.strptime(ini_d, "%Y-%m-%d") + timedelta(days=3)).strftime("%Y-%m-%d")
        prontos = {}
        print(f"\nAguardando os relatórios da faixa {ini_d}..{fim_d} (até ~30 min)...")
        for volta in range(90):                       # 90 x 20s = ~30 min
            se_faltam = [s for s in sess if s[0] not in prontos]
            if not se_faltam:
                break
            time.sleep(20)
            for sid, apelido, access in se_faltam:
                rel = achar_pronto(access, corte, alvo_ini)
                if rel:
                    prontos[sid] = (access, rel)
                    print(f"  pronto: {apelido or sid} -> {rel.get('file_name')}")
        print("=" * 55)
        for sid, apelido, access in sess:
            print(f"CONTA {apelido or sid} ({sid})")
            if sid in prontos:
                _, rel = prontos[sid]
                conteudo = baixar(access, rel["file_name"])
                if conteudo:
                    regs = ler_registros(rel["file_name"], conteudo)
                    linhas = montar_linhas(regs, sid)
                    mostrar(regs, linhas)
                    n = gravar(linhas)
                    print(f"  gravadas/atualizadas: {n} linhas.")
            else:
                print("  relatório não ficou pronto a tempo (rode de novo mais tarde).")
        print("=" * 55)
        print("✅ Fim. settlement_mp atualizada.")
        return

    # modos por conta
    for sid, apelido, access in sess:
        if MODO == "gerar":
            modo_gerar(access, sid)
        elif MODO == "baixar":
            modo_baixar(access, sid, gravar_db=True)
        else:
            sonda(access, sid)
    print("\n✅ Concluído.")


if __name__ == "__main__":
    main()
