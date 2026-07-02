"""
Robô — baixa o Relatório de Liberações do Mercado Pago e guarda em 'releases_mp'.
É a foto real do que o Mercado Pago te pagou (por pagamento), pra conciliar com as vendas.

Fluxo por conta:
  1) renova o token do ML (que também acessa o Mercado Pago)
  2) manda GERAR o relatório do período (POST) — assíncrono
  3) espera ficar pronto (consulta a lista até aparecer o arquivo)
  4) baixa o arquivo (CSV ou XLSX) e lê pela POSIÇÃO das colunas
  5) grava as linhas em 'releases_mp' (sem duplicar, via hash)

Roda no GitHub Actions, na trava 'ml-puxador' (renova o token, que é de uso único).
Período: env DIAS_MP (padrão 40 dias). Pra puxar histórico, aumente DIAS_MP.
"""

import os
import io
import csv
import json
import time
import hashlib
import requests
from datetime import datetime, timedelta, timezone
from supabase import create_client

CLIENT_ID = os.environ["ML_CLIENT_ID"]
CLIENT_SECRET = os.environ["ML_CLIENT_SECRET"]
SEED_REFRESH = os.environ.get("ML_REFRESH_TOKEN", "")
SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]
DIAS_MP = int(os.environ.get("DIAS_MP", "40"))

ML = "https://api.mercadolibre.com"
MP = "https://api.mercadopago.com"
sb = create_client(SUPABASE_URL, SUPABASE_KEY)


def renovar_token(refresh):
    r = requests.post(ML + "/oauth/token", data={
        "grant_type": "refresh_token", "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET, "refresh_token": refresh}, timeout=30)
    d = r.json()
    if "access_token" not in d:
        raise RuntimeError("Falha ao renovar token: " + str(d))
    return d


def lista_contas():
    res = sb.table("contas").select("seller_id, apelido, refresh_token").execute()
    tk = [(c["seller_id"], c.get("apelido"), c["refresh_token"])
          for c in (res.data or []) if c.get("refresh_token")]
    if not tk and SEED_REFRESH:
        tk = [(None, "(seed)", SEED_REFRESH)]
    return tk


def num(s):
    try:
        return round(float(str(s).strip() or 0), 2)
    except Exception:
        return 0.0


def parse_linha(cells, seller_id):
    c = [(x if x is not None else "") for x in cells]
    while len(c) < 15:
        c.append("")
    data = str(c[0]).strip() or None
    d = {
        "seller_id": seller_id,
        "data": data,
        "source_id": str(c[1]).strip(),
        "descricao": str(c[2]).strip(),
        "net_credit": num(c[3]),
        "net_debit": num(c[4]),
        "gross": num(c[5]),
        "mp_fee": num(c[6]),
        "taxes": num(c[7]),
        "payment_method": str(c[8]).strip(),
        "approval_date": (str(c[9]).strip() or None),
        "balance": num(c[12]),
        "payment_method_type": str(c[13]).strip(),
        "purchase_id": str(c[14]).strip(),
    }
    base = f"{seller_id}|{d['data']}|{d['source_id']}|{d['descricao']}|{d['net_credit']}|{d['net_debit']}|{d['balance']}"
    d["hash"] = hashlib.md5(base.encode()).hexdigest()
    return d


def linhas_do_arquivo(file_name, conteudo, seller_id):
    linhas = []
    if (file_name or "").lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # cabeçalho
            linhas.append(parse_linha(row, seller_id))
    else:
        texto = conteudo.decode("utf-8", errors="replace")
        rd = csv.reader(io.StringIO(texto), delimiter=";")
        for i, row in enumerate(rd):
            if i == 0 or not row:
                continue
            linhas.append(parse_linha(row, seller_id))
    return linhas


def lista_arquivos(H):
    lst = requests.get(MP + "/v1/account/release_report/list", headers=H, timeout=30).json()
    return lst if isinstance(lst, list) else []


def achar_pronto(H, corte_data, alvo_ini):
    """Acha o relatório PRONTO que cobre a faixa que a gente pediu:
       - TERMINA recentemente  (end_date  >= corte_data), e
       - COMEÇA lá atrás        (begin_date <= alvo_ini).
    O segundo teste é o que evita pegar um relatório curto (ex.: 40 dias) que
    também termina hoje mas começa há pouco tempo. Assim só aceitamos o que
    realmente cobre o período pedido (ex.: 1 ano)."""
    cand = [x for x in lista_arquivos(H)
            if x.get("file_name")
            and (x.get("end_date", "")[:10]  >= corte_data)
            and (x.get("begin_date", "")[:10] <= alvo_ini)]
    if not cand:
        return None
    # prefere a faixa mais ampla (começa mais cedo); desempata pelo mais novo
    cand.sort(key=lambda x: x.get("date_created", ""), reverse=True)
    cand.sort(key=lambda x: x.get("begin_date", ""))
    return cand[0]


def baixar_e_gravar(H, fn, sid_real):
    r = requests.get(MP + f"/v1/account/release_report/{fn}", headers=H, timeout=120)
    if r.status_code != 200:
        print(f"  falha ao baixar {fn}: {r.status_code}")
        return 0
    linhas = linhas_do_arquivo(fn, r.content, sid_real)
    n = 0
    for i in range(0, len(linhas), 200):
        lote = linhas[i:i + 200]
        try:
            sb.table("releases_mp").upsert(lote, on_conflict="hash").execute()
            n += len(lote)
        except Exception as e:
            print("  erro ao gravar lote:", str(e)[:100])
    print(f"  {fn}: {len(linhas)} movimentos lidos, {n} gravados.")
    return n


def main():
    fim = datetime.now(timezone.utc)
    ini = fim - timedelta(days=DIAS_MP)
    ini_s = ini.strftime("%Y-%m-%dT%H:%M:%SZ")
    fim_s = fim.strftime("%Y-%m-%dT%H:%M:%SZ")
    corpo = json.dumps({"begin_date": ini_s, "end_date": fim_s})

    # ---- PASSO 1: renova token e manda GERAR de todas as contas (rápido) ----
    jobs = []
    for sid, apelido, refresh in lista_contas():
        print(f"[gerar] {apelido or sid} ({sid})")
        try:
            d = renovar_token(refresh)
        except Exception as e:
            print("  não consegui renovar o token:", e)
            continue
        access = d["access_token"]
        novo = d.get("refresh_token", refresh)
        sid_real = str(d.get("user_id") or sid or "")
        if sid_real:
            sb.table("contas").upsert(
                {"seller_id": sid_real, "refresh_token": novo},
                on_conflict="seller_id").execute()
        H = {"Authorization": "Bearer " + access}
        g = requests.post(MP + "/v1/account/release_report",
                          headers={**H, "Content-Type": "application/json"},
                          data=corpo, timeout=30)
        if g.status_code not in (200, 202):
            print("  falha ao gerar:", g.status_code, g.text[:150])
            continue
        jobs.append({"sid": sid_real, "apelido": apelido, "H": H, "fn": None})

    # ---- PASSO 2: espera todos ficarem prontos e baixa (até ~30 min) ----
    corte    = (fim - timedelta(days=2)).strftime("%Y-%m-%d")   # tem que terminar "de hoje"
    alvo_ini = (ini + timedelta(days=3)).strftime("%Y-%m-%d")   # e começar lá no começo do período
    print(f"Aguardando os relatórios da faixa {ini.strftime('%Y-%m-%d')} a {fim.strftime('%Y-%m-%d')}...")
    for volta in range(90):          # 90 x 20s = ~30 min (relatório de 1 ano demora mais)
        pendentes = [j for j in jobs if not j["fn"]]
        if not pendentes:
            break
        time.sleep(20)
        for j in pendentes:
            rel = achar_pronto(j["H"], corte, alvo_ini)
            if rel:
                j["fn"] = rel.get("file_name")
                faixa = f'{rel.get("begin_date","?")[:10]} a {rel.get("end_date","?")[:10]}'
                print(f"  pronto: {j['apelido'] or j['sid']} -> {j['fn']} (faixa {faixa})")

    # ---- PASSO 3: baixa e grava ----
    print("=" * 55)
    for j in jobs:
        print(f"CONTA {j['apelido'] or j['sid']} ({j['sid']})")
        if j["fn"]:
            baixar_e_gravar(j["H"], j["fn"], j["sid"])
        else:
            print("  relatório não ficou pronto a tempo (roda de novo mais tarde).")

    print("=" * 55)
    print("✅ Fim. Tabela releases_mp atualizada.")


if __name__ == "__main__":
    main()
