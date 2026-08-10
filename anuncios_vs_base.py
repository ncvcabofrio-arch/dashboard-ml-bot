"""
Anuncios do ML que NAO estao na BaseLinker.

>>> Este arquivo mora no repositorio dashboard-ml-bot, junto dos outros. <<<

Nao reinventa nada: usa as suas proprias funcoes.
  ml_auth.obter_access()          -> token valido por conta (tabela 'contas')
  repricer_sugestoes.contas()     -> (seller_id, refresh_token) de cada conta
  repricer_sugestoes.todos_ativos -> ids dos anuncios ativos
  repricer_sugestoes.get()        -> GET com retry/429 ja tratado

So le. Nao escreve no ML, nem na BaseLinker, nem no Supabase.

-------------------------------------------------------------------
 ANTES DE RODAR
 UMA LISTA POR CONTA. Salve os MLBs que existem na BaseLinker no
 arquivo mlbs_na_base_<seller_id>.txt -- por exemplo:

     mlbs_na_base_177795203.txt    (Ponto Musical CF)

 Conta sem lista e' PULADA. Isso e' de proposito: se a lista de uma
 conta fosse usada para outra, TODOS os anuncios da outra apareceriam
 como faltando na Base -- um falso alarme de milhares de linhas.

 Secrets ja existentes no repo: SUPABASE_URL, SUPABASE_KEY,
 ML_CLIENT_ID, ML_CLIENT_SECRET.

 CUIDADO: o repricer_sugestoes usa MAX_ITENS para limitar a amostra.
 Este script zera isso, senao a varredura pararia no meio e produtos
 apareceriam como "faltando" so por nao terem sido lidos.
-------------------------------------------------------------------

 Uso:
    python3 anuncios_vs_base.py
    COM_PAUSADOS=1 python3 anuncios_vs_base.py   # inclui pausados
"""

import csv
import os
import sys
import time

import repricer_sugestoes as rec
from ml_auth import obter_access

# A amostragem do repricer nao vale aqui: precisamos do catalogo INTEIRO.
rec.MAX_ITENS = 0

COM_PAUSADOS = os.environ.get("COM_PAUSADOS", "0") == "1"

# Contas a varrer (seller_id separados por virgula). VAZIO = todas.
# Importante: so faz sentido varrer conta que tenha lista da BaseLinker,
# senao TUDO dela aparece como "faltando" -- falso alarme.
SELLERS = [s.strip() for s in os.environ.get("SELLERS", "").split(",") if s.strip()]
CAMPOS = ("id,title,status,available_quantity,price,permalink,"
          "seller_custom_field,attributes,variations")


def carregar_base(sid):
    """Lista de MLBs da BaseLinker para ESTA conta.

    Procura, nesta ordem:
        mlbs_na_base_<seller_id>.txt   <- o certo, uma lista por conta
        mlbs_na_base.txt               <- generica (aceita, mas avisa)

    Sem lista, a conta e' PULADA de proposito: comparar contra lista de
    outra conta faria todo anuncio parecer faltando.
    """
    especifico = f"mlbs_na_base_{sid}.txt"
    caminho = especifico if os.path.exists(especifico) else "mlbs_na_base.txt"
    if not os.path.exists(caminho):
        return None, None, None
    itens = set()
    with open(caminho, encoding="utf-8") as f:
        for p in f.read().replace("\n", ",").split(","):
            p = p.strip()
            if p:
                itens.add(p)
    crus = {i for i in itens if "_" not in i}   # id cru cobre todas as variacoes
    return itens, crus, caminho


def ids_pausados(sid, access):
    """O todos_ativos() so pega ativos. Pausado tambem esta no catalogo
    (tem SKU e estoque, e pode voltar a vender), entao vale conferir."""
    ids, scroll = [], None
    while True:
        p = "/users/%s/items/search?search_type=scan&limit=100&status=paused" % sid
        if scroll:
            p += "&scroll_id=" + scroll
        st, d = rec.get(p, access)
        if not isinstance(d, dict):
            break
        res = d.get("results") or []
        if not res:
            break
        ids.extend(res)
        scroll = d.get("scroll_id")
        if not scroll:
            break
    return ids


def sku_item(b):
    v = (b.get("seller_custom_field") or "").strip()
    if v:
        return v, "seller_custom_field"
    for a in (b.get("attributes") or []):
        if a.get("id") == "SELLER_SKU" and a.get("value_name"):
            return str(a["value_name"]).strip(), "SELLER_SKU"
    return "", ""


def sku_var(v, b):
    s = (v.get("seller_sku") or v.get("seller_custom_field") or "").strip()
    if s:
        return s, "variacao"
    for a in (v.get("attributes") or []):
        if a.get("id") == "SELLER_SKU" and a.get("value_name"):
            return str(a["value_name"]).strip(), "variacao.SELLER_SKU"
    return sku_item(b)


def detalhes(ids, access):
    out = []
    for i in range(0, len(ids), 20):
        lote = ids[i:i + 20]
        st, d = rec.get("/items?ids=%s&attributes=%s" % (",".join(lote), CAMPOS), access)
        if not isinstance(d, list):
            continue
        for w in d:
            b = (w or {}).get("body") or {}
            if b.get("id"):
                out.append(b)
        if (i // 20) % 20 == 0:
            print(f"    detalhes {len(out)}/{len(ids)}")
        time.sleep(0.25)
    return out


def linhas(b, sid):
    base = {"seller_id": sid, "mlb": b.get("id"), "titulo": b.get("title") or "",
            "status_ml": b.get("status") or "", "permalink": b.get("permalink") or ""}
    vs = b.get("variations") or []
    if not vs:
        sku, origem = sku_item(b)
        return [dict(base, variacao_id="", id_completo=b.get("id"), sku=sku,
                     sku_origem=origem, estoque_ml=b.get("available_quantity") or 0,
                     preco=b.get("price") or "")]
    out = []
    for v in vs:
        sku, origem = sku_var(v, b)
        out.append(dict(base, variacao_id=str(v.get("id") or ""),
                        id_completo="%s_%s" % (b.get("id"), v.get("id")),
                        sku=sku, sku_origem=origem,
                        estoque_ml=v.get("available_quantity") or 0,
                        preco=v.get("price") if v.get("price") is not None else b.get("price")))
    return out


COLS = ["seller_id", "mlb", "variacao_id", "id_completo", "titulo", "status_ml",
        "sku", "sku_origem", "estoque_ml", "preco", "na_base", "permalink"]


def main():
    todas, resumo, pulados = [], [], []
    for seller_id, refresh in rec.contas():
        if SELLERS and str(seller_id) not in SELLERS:
            continue
        access, sid, refresh = obter_access(rec.sb, seller_id, refresh)
        print(f"\n=== seller {sid} ===")

        completos, crus, caminho = carregar_base(sid)
        if completos is None:
            print(f"  PULADA: nao achei mlbs_na_base_{sid}.txt nem mlbs_na_base.txt.")
            print("  (sem a lista dessa conta, tudo apareceria como faltando)")
            pulados.append(sid)
            continue
        if caminho == "mlbs_na_base.txt" and not os.path.exists(f"mlbs_na_base_{sid}.txt"):
            print(f"  ATENCAO: usando a lista generica {caminho}.")
            print(f"  Se ela nao for desta conta, o resultado nao vale. O certo e'")
            print(f"  salvar mlbs_na_base_{sid}.txt com os MLBs desta conta.")
        print(f"  lista da Base: {len(completos)} entradas ({len(crus)} sem variacao)")
        ids, total = rec.todos_ativos(sid, access)
        print(f"  ativos: {len(ids)} (total informado pelo ML: {total})")
        if COM_PAUSADOS:
            p = ids_pausados(sid, access)
            print(f"  pausados: {len(p)}")
            ids = list(dict.fromkeys(list(ids) + p))
        dets = detalhes(ids, access)
        ls = []
        for b in dets:
            ls.extend(linhas(b, sid))
        for l in ls:
            l["na_base"] = ("SIM" if (l["id_completo"] in completos or l["mlb"] in crus)
                            else "NAO")
        todas.extend(ls)
        falt = sum(1 for l in ls if l["na_base"] == "NAO")
        ssku = sum(1 for l in ls if not l["sku"])
        resumo.append((sid, len(ids), len(ls), falt, ssku))
        print(f"  linhas: {len(ls)} | FALTANDO na Base: {falt} | sem SKU: {ssku}")

    if pulados:
        print(f"\nContas puladas por falta de lista: {pulados}")
    if not todas:
        print("\nNada coletado.")
        return

    with open("anuncios_vs_base.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=COLS, delimiter=";", extrasaction="ignore")
        w.writeheader()
        w.writerows(todas)

    falt = [l for l in todas if l["na_base"] == "NAO"]
    ssku = [l for l in todas if not l["sku"]]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = Workbook()

        def aba(nome, dados, cor="DDDDDD"):
            ws = wb.create_sheet(nome)
            ws.append(COLS)
            for c in ws[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor=cor)
            for l in dados:
                ws.append([l.get(c, "") for c in COLS])
            ws.freeze_panes = "A2"
            if dados:
                ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(dados)+1}"
            for i, w_ in enumerate([14, 16, 14, 24, 60, 10, 20, 20, 11, 11, 9, 45], 1):
                ws.column_dimensions[get_column_letter(i)].width = w_

        ws = wb.active
        ws.title = "Resumo"
        ws.append(["seller_id", "anuncios", "linhas", "FALTANDO NA BASE", "sem SKU"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in resumo:
            ws.append(list(r))
        ws.append([])
        ws.append(["TOTAL", sum(r[1] for r in resumo), len(todas), len(falt), len(ssku)])
        aba("FALTANDO NA BASE", falt, "FFC7CE")
        aba("Sem SKU", ssku, "FFE08A")
        aba("Todos", todas)
        wb.save("anuncios_vs_base.xlsx")
        print("\nOK anuncios_vs_base.xlsx e .csv")
    except ImportError:
        print("\nOK anuncios_vs_base.csv (sem openpyxl, nao gerei o xlsx)")

    print(f"   {len(todas)} linhas | FALTANDO NA BASE: {len(falt)} | sem SKU: {len(ssku)}")


if __name__ == "__main__":
    main()