"""
Teste 2 (diagnóstico) — BAIXA um relatório de liberações já pronto do Mercado Pago
e mostra as COLUNAS e algumas linhas, pra a gente ver o formato exato.
Não grava nada; só lê e imprime no log.  Roda na trava 'ml-puxador'.
"""

import os
import io
import csv
import requests
from supabase import create_client

CLIENT_ID = os.environ["ML_CLIENT_ID"]
CLIENT_SECRET = os.environ["ML_CLIENT_SECRET"]
SEED_REFRESH = os.environ.get("ML_REFRESH_TOKEN", "")
SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]

ML = "https://api.mercadolibre.com"
MP = "https://api.mercadopago.com"
sb = create_client(SUPABASE_URL, SUPABASE_KEY)


def renovar_token(refresh):
    r = requests.post(ML + "/oauth/token", data={
        "grant_type": "refresh_token", "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET, "refresh_token": refresh}, timeout=30)
    d = r.json()
    if "access_token" not in d:
        raise RuntimeError("Falha ao renovar token: " + str(d))
    return d


def lista_contas():
    res = sb.table("contas").select("seller_id, apelido, refresh_token").execute()
    tk = [(c["seller_id"], c.get("apelido"), c["refresh_token"])
          for c in (res.data or []) if c.get("refresh_token")]
    if not tk and SEED_REFRESH:
        tk = [(None, "(seed)", SEED_REFRESH)]
    return tk


def mostra_xlsx(conteudo):
    try:
        from openpyxl import load_workbook
    except Exception:
        print("   (openpyxl não instalado — não consegui abrir o xlsx)")
        return
    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    linhas = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        linhas.append(row)
        if i >= 4:
            break
    if linhas:
        print("   COLUNAS:", list(linhas[0]))
        for r in linhas[1:]:
            print("   linha:", list(r))


def mostra_csv(texto):
    rd = list(csv.reader(io.StringIO(texto)))
    if rd:
        print("   COLUNAS:", rd[0])
        for r in rd[1:4]:
            print("   linha:", r)


def main():
    for sid, apelido, refresh in lista_contas():
        print("=" * 60)
        print(f"CONTA {apelido or sid} ({sid})")
        try:
            d = renovar_token(refresh)
        except Exception as e:
            print("  não consegui renovar o token:", e)
            continue
        access = d["access_token"]
        novo = d.get("refresh_token", refresh)
        if sid:
            sb.table("contas").upsert(
                {"seller_id": str(d.get("user_id") or sid), "refresh_token": novo},
                on_conflict="seller_id").execute()
        H = {"Authorization": "Bearer " + access}

        lst = requests.get(MP + "/v1/account/release_report/list", headers=H, timeout=30).json()
        if not isinstance(lst, list) or not lst:
            print("  (sem relatório pronto ainda — gere um e rode de novo em uns minutos)")
            continue
        # pega o mais recente
        rel = sorted(lst, key=lambda x: x.get("date_created", ""), reverse=True)[0]
        fn = rel.get("file_name")
        print(f"  Baixando: {fn}")
        r = requests.get(MP + f"/v1/account/release_report/{fn}", headers=H, timeout=60)
        print(f"  download HTTP {r.status_code}, {len(r.content)} bytes")
        if r.status_code != 200:
            print("   ", r.text[:200])
            continue
        if (fn or "").lower().endswith(".xlsx"):
            mostra_xlsx(r.content)
        else:
            mostra_csv(r.text)

    print("=" * 60)
    print("Me manda as COLUNAS que apareceram — com elas eu monto a tabela e a conciliação.")


if __name__ == "__main__":
    main()
